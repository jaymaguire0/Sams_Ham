import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
import logging

# Setup logging
logging.basicConfig(
    filename="update_log.txt",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

def update_ts_file(file_path, project_name, project_number):
    """Update Technical Submittal (TS) file."""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        ws["B5"] = project_name if project_name else ""
        ws["B6"] = project_number if project_number else ""
        wb.save(file_path)
        logging.info(f"Updated TS file: {file_path}")
    except Exception as e:
        logging.error(f"Error updating TS file {file_path}: {e}")

def update_es_file(file_path, project_name, project_number, issued_for):
    """Update Equipment Schedule (ES) file by searching for specific keywords."""
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # Exact keyword mapping
        keywords = {
            "Project No.": project_number,
            "Project Name": project_name,
            "Issued For": issued_for
        }

        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if cell.value in keywords:  # If the cell text matches one of our keywords
                    target_value = keywords[cell.value]
                    ws.cell(
                        row=cell.row,
                        column=cell.column + 1,
                        value=target_value if target_value else ""  # Write into the next cell
                    )

        wb.save(file_path)
        logging.info(f"Updated ES file: {file_path}")
    except Exception as e:
        logging.error(f"Error updating ES file {file_path}: {e}")

def run_update(folder_path, project_name, project_number, issued_for, update_ts, update_es, progress_var, root):
    """Run updates on selected files with progress bar."""
    try:
        # Collect all files recursively
        files = []
        for root_dir, _, filenames in os.walk(folder_path):
            for f in filenames:
                files.append(os.path.join(root_dir, f))

        # Filter tasks
        tasks = []
        if update_ts:
            tasks.extend([f for f in files if os.path.basename(f).startswith("MST-TS") and f.endswith(".xlsx")])
        if update_es:
            tasks.extend([f for f in files if os.path.basename(f).startswith("ES-") and f.endswith(".xlsx")])

        total = len(tasks)
        if total == 0:
            messagebox.showwarning("No Files", "No matching files found to update.")
            return

        for i, file_path in enumerate(tasks, start=1):
            if os.path.basename(file_path).startswith("MST-TS") and update_ts:
                update_ts_file(file_path, project_name, project_number)
            elif os.path.basename(file_path).startswith("ES-") and update_es:
                update_es_file(file_path, project_name, project_number, issued_for)
            progress_var.set(int((i / total) * 100))
            root.update_idletasks()

        messagebox.showinfo("Success", "Files updated successfully!")
    except Exception as e:
        logging.error(f"Error running update: {e}")
        messagebox.showerror("Error", f"An error occurred: {e}")

def run_gui():
    """Main GUI."""
    root = tk.Tk()
    root.title("Update Project Info V3")

    # Disable resizing
    root.resizable(False, False)

    # Center columns
    root.grid_columnconfigure(0, weight=1)
    root.grid_columnconfigure(1, weight=1)
    root.grid_columnconfigure(2, weight=1)

    # Variables
    project_name_var = tk.StringVar()
    project_number_var = tk.StringVar()
    issued_for_var = tk.StringVar()
    folder_path_var = tk.StringVar()
    update_ts_var = tk.BooleanVar(value=True)
    update_es_var = tk.BooleanVar(value=False)

    # Folder selection
    tk.Label(root, text="Select Folder:", anchor="center", justify="center").grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
    tk.Entry(root, textvariable=folder_path_var, width=50, justify="center").grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
    tk.Button(root, text="Browse", command=lambda: folder_path_var.set(filedialog.askdirectory())).grid(row=0, column=2, padx=5, pady=5, sticky="nsew")

    # Checkboxes
    tk.Checkbutton(root, text="Technical Submittal", variable=update_ts_var, anchor="center", justify="center").grid(row=1, column=0, columnspan=3, sticky="nsew")
    tk.Checkbutton(root, text="Equipment Schedule", variable=update_es_var, anchor="center", justify="center").grid(row=2, column=0, columnspan=3, sticky="nsew")

    # Inputs
    tk.Label(root, text="Project Name:", anchor="center", justify="center").grid(row=3, column=0, padx=5, pady=5, sticky="nsew")
    tk.Entry(root, textvariable=project_name_var, width=40, justify="center").grid(row=3, column=1, columnspan=2, padx=5, pady=5, sticky="nsew")

    tk.Label(root, text="Project Number:", anchor="center", justify="center").grid(row=4, column=0, padx=5, pady=5, sticky="nsew")
    tk.Entry(root, textvariable=project_number_var, width=40, justify="center").grid(row=4, column=1, columnspan=2, padx=5, pady=5, sticky="nsew")

    tk.Label(root, text="Issued For:", anchor="center", justify="center").grid(row=5, column=0, padx=5, pady=5, sticky="nsew")
    issued_for_entry = tk.Entry(root, textvariable=issued_for_var, width=40, justify="center", state="disabled")
    issued_for_entry.grid(row=5, column=1, columnspan=2, padx=5, pady=5, sticky="nsew")

    def toggle_issued_for():
        if update_es_var.get():
            issued_for_entry.config(state="normal")
        else:
            issued_for_entry.config(state="disabled")
            issued_for_var.set("")

    update_es_var.trace_add("write", lambda *args: toggle_issued_for())

    # Progress bar
    progress_var = tk.IntVar()
    progress = ttk.Progressbar(root, variable=progress_var, maximum=100)
    progress.grid(row=6, column=0, columnspan=3, padx=10, pady=10, sticky="nsew")

    # Update button
    tk.Button(
        root,
        text="Update Files",
        command=lambda: run_update(
            folder_path_var.get(),
            project_name_var.get(),
            project_number_var.get(),
            issued_for_var.get(),
            update_ts_var.get(),
            update_es_var.get(),
            progress_var,
            root
        )
    ).grid(row=7, column=0, columnspan=3, pady=10, sticky="nsew")

    # Notes & Warnings at the bottom
    notes = [
        "Important Notes:",
        "- Do NOT run this tool if the Excel files are already open (on your PC or another's).",
        "- Ensure the folder contains the correct MST-TS and/or ES files before running.",
        "- Empty fields will clear the corresponding Excel cells.",
        "- If 'Equipment Schedule' is selected, 'Issued For' must be provided (or left blank intentionally).",
        "- Always keep a backup of your files before bulk updates.",
        "",  # blank line before footer
        "P.S. SAM IS A HAM"
    ]

    for i, note in enumerate(notes, start=8):
        if note.startswith("Important Notes:"):
            font_style = ("Arial", 10, "bold", "underline")
        elif note.startswith("P.S."):
            font_style = ("Arial", 10, "bold")
        else:
            font_style = ("Arial", 10)

        tk.Label(root, text=note, anchor="center", justify="center", font=font_style).grid(
            row=i, column=0, columnspan=3, padx=10, sticky="nsew"
        )

    root.mainloop()

if __name__ == "__main__":
    run_gui()
