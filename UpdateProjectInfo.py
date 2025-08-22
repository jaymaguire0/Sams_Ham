import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from datetime import datetime

def run_update():
    project_name = entry_name.get().strip()
    project_number = entry_number.get().strip()
    root_folder = entry_folder.get().strip()

    if not project_name or not project_number or not root_folder:
        messagebox.showerror("Missing Info", "Please fill in all fields before running.")
        return

    # Collect all matching files first (to know progress total)
    matching_files = []
    for folder, subfolders, files in os.walk(root_folder):
        for file in files:
            if file.startswith("MST-TS") and file.endswith(".xlsx"):
                matching_files.append(os.path.join(folder, file))

    total_files = len(matching_files)
    if total_files == 0:
        messagebox.showinfo("No Files", "No matching MST-TS Excel files found.")
        return

    # Prepare log file (in same folder as exe/script)
    log_filename = f"UpdateLog_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    log_path = os.path.join(os.getcwd(), log_filename)
    with open(log_path, "w", encoding="utf-8") as log:
        log.write(f"Project Info Updater Log - {datetime.now()}\n")
        log.write(f"Project Name: {project_name}\n")
        log.write(f"Project Number: {project_number}\n")
        log.write(f"Root Folder: {root_folder}\n")
        log.write("=" * 50 + "\n\n")

        updated_files = 0
        skipped_files = []

        # Reset and configure progress bar
        progress["maximum"] = total_files
        progress["value"] = 0
        root.update_idletasks()

        for i, file_path in enumerate(matching_files, start=1):
            try:
                wb = load_workbook(file_path)
                ws = wb.active  # change to wb["SheetName"] if needed
                ws["B5"] = project_name
                ws["B6"] = project_number
                wb.save(file_path)

                updated_files += 1
                log.write(f"SUCCESS: {file_path}\n")
            except PermissionError:
                skipped_files.append(file_path)
                log.write(f"SKIPPED (file in use/locked): {file_path}\n")
            except Exception as e:
                log.write(f"FAILED: {file_path} - {e}\n")

            # Update progress bar
            progress["value"] = i
            root.update_idletasks()

        # Final summary
        log.write("\n" + "=" * 50 + "\n")
        log.write(f"Updated {updated_files} of {total_files} files.\n")
        if skipped_files:
            log.write(f"Skipped {len(skipped_files)} files (in use):\n")
            for f in skipped_files:
                log.write(f"   {f}\n")
        log.write(f"\nLog file saved at: {log_path}\n")

    summary = f"✅ Updated {updated_files}/{total_files} files.\nLog saved to:\n{log_path}"
    if skipped_files:
        summary += f"\n⚠️ {len(skipped_files)} files skipped (open or locked)."
    messagebox.showinfo("Process Complete", summary)


def browse_folder():
    folder_selected = filedialog.askdirectory()
    if folder_selected:
        entry_folder.delete(0, tk.END)
        entry_folder.insert(0, folder_selected)


# === GUI Setup ===
root = tk.Tk()
root.title("Excel Project Info Updater")
root.geometry("550x320")

# Labels & Inputs
tk.Label(root, text="Project Name:", anchor="w").pack(fill="x", padx=10, pady=5)
entry_name = tk.Entry(root)
entry_name.pack(fill="x", padx=10)

tk.Label(root, text="Project Number:", anchor="w").pack(fill="x", padx=10, pady=5)
entry_number = tk.Entry(root)
entry_number.pack(fill="x", padx=10)

tk.Label(root, text="Root Folder (contains all subfolders):", anchor="w").pack(fill="x", padx=10, pady=5)
frame_folder = tk.Frame(root)
frame_folder.pack(fill="x", padx=10)
entry_folder = tk.Entry(frame_folder)
entry_folder.pack(side="left", fill="x", expand=True)
tk.Button(frame_folder, text="Browse...", command=browse_folder).pack(side="left")

# Progress Bar
progress = ttk.Progressbar(root, orient="horizontal", length=500, mode="determinate")
progress.pack(pady=10, padx=10)

# Run Button
tk.Button(root, text="Run Update", command=run_update, bg="green", fg="white").pack(pady=10)

# Instructions
tk.Label(root, text="This updates B5 (Project Name) and B6 (Project Number)\n"
         "in all MST-TS Excel files inside the selected folder.\n"
         "A log file will be saved in the same folder as this program.",
         fg="gray").pack(pady=5)

root.mainloop()
